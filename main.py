# -*- coding: utf-8 -*-
"""
Serviço PLJ — Gerador de Work Order (padrão PLJ, com fotos traduzidas em PT).

Fluxo:
  POST /gerar {"num": "4807"}  + header Authorization: Bearer <token do usuário Supabase>
    -> confere se o usuário é admin (tabela app_users no Supabase)
    -> acha a estimate no JobNimbus pelo número
    -> baixa o PDF do SumoQuote (attachment_id)
    -> pdfimages extrai as fotos (filtra as pequenas: logo/ícone)
    -> Claude lê o PDF + as fotos numeradas e devolve JSON estruturado em PT
    -> reportlab monta o PDF no padrão PLJ
    -> sobe no Storage do Supabase (bucket workorders) e devolve a URL pública
"""
import os
import io
import time
import json
import glob
import base64
import tempfile
import subprocess

import logging
import requests

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("plj")
from fastapi import FastAPI, Header, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from PIL import Image as PILImage

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Image,
                                Table, TableStyle, KeepTogether, HRFlowable,
                                PageBreak)

# ---------------------------------------------------------------- config
JOBNIMBUS_KEY = os.environ.get("JOBNIMBUS_KEY", "")
ANTHROPIC_KEY = os.environ.get("ANTHROPIC_KEY", "")
SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
SUPABASE_ANON_KEY = os.environ.get("SUPABASE_ANON_KEY", "")
BUCKET = os.environ.get("WO_BUCKET", "workorders")
MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-6")

JN_BASE = "https://app.jobnimbus.com/api1"
MAX_FOTOS = 24          # teto de fotos mandadas pro Claude
FOTO_MAX_PX = 1400      # redimensiona antes de mandar pra API

app = FastAPI(title="PLJ Work Order generator")
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_methods=["*"], allow_headers=["*"])


class Req(BaseModel):
    num: str


# ---------------------------------------------------------------- estilos
YEL = colors.HexColor("#F2C200")
DARK = colors.HexColor("#1F1F1F")
GREY = colors.HexColor("#555555")

S_TITULO = ParagraphStyle("titulo", fontName="Helvetica-Bold", fontSize=16,
                          leading=19, textColor=DARK, alignment=TA_CENTER,
                          spaceAfter=2)
S_SUB = ParagraphStyle("sub", fontName="Helvetica", fontSize=9.5, leading=12,
                       textColor=GREY, alignment=TA_CENTER, spaceAfter=6)
S_SECAO = ParagraphStyle("secao", fontName="Helvetica-Bold", fontSize=12,
                         leading=15, textColor=DARK, spaceBefore=10,
                         spaceAfter=4)
S_NOTA = ParagraphStyle("nota", fontName="Helvetica-Oblique", fontSize=9.5,
                        leading=12.5, textColor=GREY, spaceAfter=6)
S_ETAPA = ParagraphStyle("etapa", fontName="Helvetica", fontSize=10,
                         leading=13.5, spaceAfter=5)
S_BULLET = ParagraphStyle("bullet", fontName="Helvetica", fontSize=10,
                          leading=13.5, leftIndent=12, bulletIndent=2,
                          spaceAfter=3)
S_FOTOT = ParagraphStyle("fotot", fontName="Helvetica-Bold", fontSize=10.5,
                         leading=13, textColor=DARK, spaceBefore=6,
                         spaceAfter=2)
S_FOTOD = ParagraphStyle("fotod", fontName="Helvetica", fontSize=9.5,
                         leading=12.5, spaceAfter=10)
S_CELL = ParagraphStyle("cell", fontName="Helvetica", fontSize=9.5, leading=12)
S_CELLB = ParagraphStyle("cellb", fontName="Helvetica-Bold", fontSize=9.5,
                         leading=12)


def _esc(t):
    return (str(t or "").replace("&", "&amp;")
            .replace("<", "&lt;").replace(">", "&gt;"))


# ---------------------------------------------------------------- jobnimbus
def jn_headers():
    return {"Authorization": "Bearer " + JOBNIMBUS_KEY,
            "Accept": "application/json"}


def find_estimate(num):
    flt = json.dumps({"must": [{"term": {"number": str(num)}}]})
    r = requests.get(f"{JN_BASE}/estimates", headers=jn_headers(),
                     params={"filter": flt}, timeout=60)
    if not r.ok:
        raise HTTPException(502, f"JobNimbus falhou ({r.status_code}).")
    results = (r.json() or {}).get("results") or []
    if not results:
        raise HTTPException(404, f"Estimate {num} não encontrada no JobNimbus.")
    return results[0]


def jn_contato(est):
    """Busca o contato ligado à estimate — é onde mora o endereço bom."""
    c = next((r for r in (est.get("related") or [])
              if r.get("type") == "contact"), None)
    if not c or not c.get("id"):
        return {}
    try:
        r = requests.get(f"{JN_BASE}/contacts/{c['id']}",
                         headers=jn_headers(), timeout=60)
        return r.json() if r.ok else {}
    except Exception:
        return {}


def dados_cliente(est):
    """Nome e endereço em partes, com a estimate como reserva."""
    ct = jn_contato(est)
    rel = next((r for r in (est.get("related") or [])
                if r.get("type") == "contact"), None)
    nome = (ct.get("display_name") or ct.get("name")
            or (rel or {}).get("name") or "").strip()
    def pega(k):
        return str(ct.get(k) or est.get(k) or "").strip()
    rua = " ".join(x for x in [pega("address_line1"),
                               pega("address_line2")] if x).strip()
    cid = pega("city")
    est_uf = pega("state_text")
    cep = pega("zip")
    completo = ", ".join(x for x in
                         [rua, cid, " ".join(y for y in [est_uf, cep] if y)]
                         if x)
    return {"cliente": nome, "rua": rua, "cidade": cid,
            "estado": est_uf, "cep": cep, "endereco": completo}


def jn_link(est):
    """Monta o endereço da estimate no site do JobNimbus.

    O site localiza a estimate pelo jnid; o trecho do job é ignorado,
    então serve qualquer id relacionado (usamos o do job quando existe).
    """
    est_id = str(est.get("jnid") or est.get("guid") or "").strip().lower()
    if not est_id:
        return ""
    rel = est.get("related") or []
    job = next((r for r in rel if r.get("type") == "job"), None)
    jobid = str((job or {}).get("id") or "").strip()
    if not jobid:
        jobid = str((rel[0] if rel else {}).get("id") or "").strip() or "0"
    return ("https://app.jobnimbus.com/job/" + jobid +
            "/estimates/" + est_id + "/view")


def download_pdf(attachment_id):
    r = requests.get(f"{JN_BASE}/files/{attachment_id}",
                     headers=jn_headers(), timeout=180)
    if not r.ok or not r.content[:4] == b"%PDF":
        raise HTTPException(502, "Não consegui baixar o PDF da estimate.")
    return r.content


# ---------------------------------------------------------------- claude
PROMPT = """Você escreve a ORDEM DE SERVIÇO da PLJ Carpentry (Cape Cod, MA) para a
equipe de campo ler no celular, em pé no canteiro. Recebe o PDF de uma estimate do
SumoQuote (em inglês) e as fotos extraídas dele, numeradas.

COMO ESCREVER — isto é o mais importante:
- PORTUGUÊS SIMPLES, do dia a dia da obra. Nada de linguagem formal ou técnica.
- Frases CURTAS: no máximo 12 palavras. Uma ação por frase.
- Verbo direto: "Trocar", "Instalar", "Tirar", "Vedar". Nunca "proceder à instalação",
  "efetuar a remoção", "realizar a substituição".
- Palavra simples sempre que der: "tirar" e não "remover"; "colocar" e não "posicionar";
  "medir" e não "aferir".
- CORTE O ÓBVIO. Não escreva o que qualquer carpinteiro já sabe (medir antes de cortar,
  usar EPI, limpar no fim, trabalhar com cuidado, proteger o piso). Só entra o que é
  específico DESTE job.
- Sem enrolação, sem introdução, sem repetir a mesma informação em dois lugares.
- Se uma etapa não muda nada na prática, não escreva ela.
- Nunca inclua preços, valores ou totais.

REGRAS DE CONTEÚDO:
- Escreva tudo em português (só nomes próprios, marcas e códigos ficam como estão).
- Nas fotos, DIGA A COR DA MARCAÇÃO quando existir ("o que está marcado em vermelho").
- Título da etapa: 1 a 3 palavras (o local ou a peça). Ex.: "PORTA DA FRENTE".
- Descrição da etapa: 1 frase curta.
- Descrição da foto: 1 frase curta dizendo o que fazer ali.
- "atencao": no MÁXIMO 4 itens, só o que causa retrabalho, erro caro ou atraso
  (divergência de marca/medida, esperar permit, material que ainda não chegou).
  Se não tiver nada de verdade, devolva lista vazia.

O CABEÇALHO — só estes campos, nada mais:
- "endereco": endereço do imóvel, copiado do PDF. Se não achar, devolva "".
- "vendedor": nome do vendedor/sales rep, copiado do PDF. Se não achar, devolva "".
- "servicos": lista curta com os TIPOS de serviço, 2 a 5 palavras cada.
  Ex.: ["Troca de 6 janelas", "Trim de PVC"]. Só o tipo, sem detalhe nem medida.
- NÃO INVENTE NADA. Se a informação não está no PDF, devolva "" ou lista vazia.
  Não deduza, não estime, não complete com o que "provavelmente" é.
- NÃO coloque no cabeçalho: forma de pagamento, valores, datas, número do contrato,
  especificação técnica completa, nem repetir o nome do cliente.
- DESCARTE (manter=false) fotos que sejam logo, capa, certificado de seguro,
  certificação de fabricante, tabela de preços, assinatura ou página só de texto.
  Fique só com fotos reais do imóvel.

EXEMPLOS DO TOM CERTO:
  RUIM: "Proceder à remoção do trim de madeira existente ao redor do vão da porta
         frontal e efetuar a instalação de novo trim em PVC."
  BOM:  "Tirar o trim de madeira da porta e colocar trim de PVC."
  RUIM: "Verificar cuidadosamente as medidas antes de realizar os cortes."
  BOM:  (não escrever — é óbvio)

Responda SOMENTE com um objeto JSON válido, sem markdown, sem crases, no formato:
{
  "titulo": "TRIM E JANELAS",
  "endereco": "19 Janet Street, Plymouth, MA 02360",
  "vendedor": "Rob Silva",
  "servicos": ["Troca de 6 janelas", "Trim de PVC"],
  "partes": [
    {"titulo": "PARTE A — TRIM DE PVC",
     "nota": "Trocar só o que está marcado em vermelho.",
     "etapas": [{"titulo": "PORTA DA FRENTE",
                 "descricao": "Tirar o trim de madeira e colocar trim de PVC."}]}
  ],
  "atencao": ["Não começar antes do permit sair."],
  "fotos": [{"foto": 1, "manter": true, "titulo": "Frente da casa",
             "descricao": "Trocar as 2 janelas marcadas em amarelo."}]
}
O campo "foto" é o número da foto conforme enviada. Inclua TODAS as fotos na lista,
com manter=true ou manter=false."""


def _img_block(path):
    im = PILImage.open(path).convert("RGB")
    if max(im.size) > FOTO_MAX_PX:
        r = FOTO_MAX_PX / float(max(im.size))
        im = im.resize((int(im.size[0] * r), int(im.size[1] * r)))
    buf = io.BytesIO()
    im.save(buf, format="JPEG", quality=82)
    return {"type": "image",
            "source": {"type": "base64", "media_type": "image/jpeg",
                       "data": base64.b64encode(buf.getvalue()).decode()}}


def claude_analyze(pdf_bytes, photos, cliente, endereco, total):
    content = [{
        "type": "document",
        "source": {"type": "base64", "media_type": "application/pdf",
                   "data": base64.b64encode(pdf_bytes).decode()}
    }]
    for i, p in enumerate(photos[:MAX_FOTOS], 1):
        content.append({"type": "text", "text": f"FOTO {i}:"})
        content.append(_img_block(p))
    content.append({"type": "text", "text":
                    f"Cliente: {cliente}\nEndereço: {endereco}\n\n{PROMPT}"})

    r = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={"x-api-key": ANTHROPIC_KEY,
                 "anthropic-version": "2023-06-01",
                 "content-type": "application/json"},
        json={"model": MODEL, "max_tokens": 8000,
              "messages": [{"role": "user", "content": content}]},
        timeout=600)
    if not r.ok:
        raise HTTPException(502, f"Claude falhou ({r.status_code}): {r.text[:300]}")
    txt = "".join(b.get("text", "") for b in r.json().get("content", [])).strip()
    if txt.startswith("```"):
        txt = txt.split("```")[1]
        if txt.startswith("json"):
            txt = txt[4:]
    a = txt.find("{")
    b = txt.rfind("}")
    if a == -1 or b == -1:
        raise HTTPException(502, "Claude não devolveu JSON.")
    return json.loads(txt[a:b + 1])


# ---------------------------------------------------------------- pdf
def _linha_amarela():
    return HRFlowable(width="100%", thickness=2.2, color=YEL,
                      spaceBefore=3, spaceAfter=8)


def build_pdf(out_path, wo, cliente, endereco, vendedor, servicos, photos):
    doc = SimpleDocTemplate(out_path, pagesize=letter,
                            leftMargin=0.8 * inch, rightMargin=0.8 * inch,
                            topMargin=0.7 * inch, bottomMargin=0.7 * inch,
                            title="Ordem de Serviço — PLJ Carpentry")
    F = []
    titulo = _esc(wo.get("titulo", "")).upper()
    F.append(Paragraph(f"ORDEM DE SERVIÇO — {titulo}", S_TITULO))
    F.append(Paragraph("PLJ CARPENTRY &amp; REMODELING", S_SUB))
    F.append(_linha_amarela())

    if isinstance(servicos, (list, tuple)):
        servicos = " · ".join(str(s) for s in servicos if s)
    linhas = [["CLIENTE", cliente], ["ENDEREÇO", endereco],
              ["VENDEDOR", vendedor], ["SERVIÇOS", servicos]]
    linhas = [[k, v] for k, v in linhas if str(v or "").strip()]
    dados = [[Paragraph(_esc(k), S_CELLB), Paragraph(_esc(v), S_CELL)]
             for k, v in linhas]
    t = Table(dados, colWidths=[1.35 * inch, 5.55 * inch])
    t.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#CCCCCC")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F5F5F5")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    F.append(t)
    F.append(Spacer(1, 10))

    for parte in (wo.get("partes") or []):
        F.append(Paragraph(_esc(parte.get("titulo", "")).upper(), S_SECAO))
        if parte.get("nota"):
            F.append(Paragraph(_esc(parte["nota"]), S_NOTA))
        for i, et in enumerate(parte.get("etapas") or [], 1):
            tit = _esc(et.get("titulo", ""))
            des = _esc(et.get("descricao", ""))
            F.append(Paragraph(f"<b>{i}. {tit}</b> — {des}", S_ETAPA))

    if wo.get("atencao"):
        F.append(Paragraph("ATENÇÃO", S_SECAO))
        for b in wo["atencao"]:
            F.append(Paragraph(_esc(b), S_BULLET, bulletText="•"))

    if photos:
        F.append(PageBreak())
        F.append(Paragraph("FOTOS DO LOCAL", S_SECAO))
        F.append(_linha_amarela())
        for path, tit, des in photos:
            try:
                iw, ih = PILImage.open(path).size
            except Exception:
                continue
            w = 4.9 * inch
            h = w * (ih / float(iw))
            if h > 5.4 * inch:
                h = 5.4 * inch
                w = h * (iw / float(ih))
            bloco = [Image(path, width=w, height=h)]
            if tit:
                bloco.append(Paragraph(_esc(tit), S_FOTOT))
            if des:
                bloco.append(Paragraph(_esc(des), S_FOTOD))
            F.append(KeepTogether(bloco))

    doc.build(F)
    return out_path


# ---------------------------------------------------------------- supabase
def quem_e(authorization):
    """Devolve (email, papel) de quem está chamando."""
    if not authorization:
        raise HTTPException(401, "Sem token.")
    tok = authorization.replace("Bearer ", "").strip()
    u = requests.get(f"{SUPABASE_URL}/auth/v1/user",
                     headers={"Authorization": "Bearer " + tok,
                              "apikey": SUPABASE_ANON_KEY}, timeout=30)
    if not u.ok:
        raise HTTPException(401, "Token inválido.")
    email = (u.json() or {}).get("email")
    q = requests.get(f"{SUPABASE_URL}/rest/v1/app_users",
                     params={"email": f"eq.{email}", "select": "role,name"},
                     headers={"Authorization": "Bearer " + SUPABASE_SERVICE_KEY,
                              "apikey": SUPABASE_SERVICE_KEY}, timeout=30)
    linhas = q.json() if q.ok else []
    if not linhas:
        raise HTTPException(403, "Usuário sem permissão.")
    return email, (linhas[0].get("role") or ""), (linhas[0].get("name") or "")


def check_admin(authorization):
    if not authorization:
        raise HTTPException(401, "Sem token.")
    tok = authorization.replace("Bearer ", "").strip()
    u = requests.get(f"{SUPABASE_URL}/auth/v1/user",
                     headers={"Authorization": "Bearer " + tok,
                              "apikey": SUPABASE_ANON_KEY}, timeout=30)
    if not u.ok:
        raise HTTPException(401, "Token inválido.")
    email = (u.json() or {}).get("email")
    q = requests.get(f"{SUPABASE_URL}/rest/v1/app_users",
                     params={"email": f"eq.{email}", "select": "role"},
                     headers={"Authorization": "Bearer " + SUPABASE_SERVICE_KEY,
                              "apikey": SUPABASE_SERVICE_KEY}, timeout=30)
    rows = q.json() if q.ok else []
    if not rows or rows[0].get("role") != "admin":
        raise HTTPException(403, "Só admin pode gerar Work Order.")
    return email


def upload_pdf(local_path, num):
    nome = f"WO_{num}_{os.urandom(3).hex()}.pdf"
    with open(local_path, "rb") as f:
        r = requests.post(
            f"{SUPABASE_URL}/storage/v1/object/{BUCKET}/{nome}",
            headers={"Authorization": "Bearer " + SUPABASE_SERVICE_KEY,
                     "apikey": SUPABASE_SERVICE_KEY,
                     "Content-Type": "application/pdf",
                     "x-upsert": "true"},
            data=f.read(), timeout=180)
    if not r.ok:
        raise HTTPException(502, f"Falha ao subir no Storage: {r.text[:200]}")
    return f"{SUPABASE_URL}/storage/v1/object/public/{BUCKET}/{nome}"


def extrair_fotos(pdf_path, tmpdir):
    subprocess.run(["pdfimages", "-all", "-p", pdf_path,
                    os.path.join(tmpdir, "img")], check=False)
    out = []
    for p in sorted(glob.glob(os.path.join(tmpdir, "img*"))):
        try:
            w, h = PILImage.open(p).size
            if w >= 200 and h >= 150:
                if not p.lower().endswith(".png"):
                    np_ = p + ".png"
                    PILImage.open(p).convert("RGB").save(np_)
                    p = np_
                out.append(p)
        except Exception:
            pass
    return out


# ---------------------------------------------------------------- rotas
@app.get("/")
def health():
    return {"ok": True, "service": "PLJ Work Order generator"}


@app.post("/link")
def link(req: Req, authorization: str = Header(default="")):
    """Só devolve o endereço da estimate no JobNimbus. Rápido, sem gerar PDF."""
    check_admin(authorization)
    est = find_estimate(req.num)
    d = dados_cliente(est)
    return {"ok": True, "num": req.num, "jn_url": jn_link(est),
            "cliente": d["cliente"], "rua": d["rua"], "cidade": d["cidade"],
            "estado": d["estado"], "cep": d["cep"], "endereco": d["endereco"],
            "total": est.get("total"), "vendedor": est.get("sales_rep_name", "")}


@app.post("/gerar")
def gerar(req: Req, authorization: str = Header(default="")):
    check_admin(authorization)
    est = find_estimate(req.num)
    att = est.get("attachment_id")
    if not att:
        raise HTTPException(404, "Estimate sem PDF anexo.")

    d = dados_cliente(est)
    cliente = d["cliente"]
    endereco = d["endereco"]
    vendedor = str(est.get("sales_rep_name", "") or "").strip()
    pdf_bytes = download_pdf(att)

    with tempfile.TemporaryDirectory() as tmp:
        src = os.path.join(tmp, "src.pdf")
        with open(src, "wb") as f:
            f.write(pdf_bytes)

        fotos = extrair_fotos(src, tmp)
        analise = claude_analyze(pdf_bytes, fotos, cliente, endereco,
                                 est.get("total"))

        info = {int(f.get("foto", 0)): f for f in (analise.get("fotos") or [])}
        wo_photos = []
        for i, p in enumerate(fotos[:MAX_FOTOS], 1):
            dfoto = info.get(i)
            if dfoto and dfoto.get("manter"):
                wo_photos.append((p, dfoto.get("titulo", ""),
                                  dfoto.get("descricao", "")))

        out = os.path.join(tmp, "out.pdf")
        end_final = endereco or str(analise.get("endereco", "") or "").strip()
        vend_final = vendedor or str(analise.get("vendedor", "") or "").strip()
        build_pdf(out, analise, cliente, end_final, vend_final,
                  analise.get("servicos") or "", wo_photos)
        url = upload_pdf(out, req.num)

    return {"ok": True, "cliente": cliente, "url": url,
            "jn_url": jn_link(est),
            "titulo": analise.get("titulo", ""),
            "fotos_incluidas": len(wo_photos),
            "fotos_extraidas": len(fotos)}


# =====================================================================
# QUICKBOOKS — OAuth, tokens e financeiro por projeto
# =====================================================================
import re as _re
import secrets as _secrets
from datetime import date as _date
from fastapi.responses import RedirectResponse, HTMLResponse

QB_ID = os.environ.get("QB_CLIENT_ID", "")
QB_SECRET = os.environ.get("QB_CLIENT_SECRET", "")
SERVICE_URL = os.environ.get("SERVICE_URL", "").rstrip("/")
QB_REDIRECT = SERVICE_URL + "/qb/callback"
QB_DISCOVERY = "https://developer.api.intuit.com/.well-known/openid_configuration"
QB_AUTH_PADRAO = "https://appcenter.intuit.com/connect/oauth2"
QB_TOKEN_PADRAO = "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer"
QB_API = "https://quickbooks.api.intuit.com/v3/company"
QB_MINOR = "75"
_qb_estado = {}
_qb_cache = {"access": None, "exp": 0, "realm": None}
_qb_endp = {"auth": None, "token": None, "quando": 0}


def qb_endpoints():
    """Pega os endereços no discovery document da Intuit (com reserva fixa)."""
    agora = time.time()
    if _qb_endp["auth"] and (agora - _qb_endp["quando"]) < 86400:
        return _qb_endp["auth"], _qb_endp["token"]
    a, t = QB_AUTH_PADRAO, QB_TOKEN_PADRAO
    try:
        r = requests.get(QB_DISCOVERY, timeout=20)
        if r.ok:
            j = r.json()
            a = j.get("authorization_endpoint") or a
            t = j.get("token_endpoint") or t
    except Exception:
        pass
    _qb_endp.update({"auth": a, "token": t, "quando": agora})
    return a, t


# ---------------------------------------------------------- tokens
def qb_sb_headers():
    return {"Authorization": "Bearer " + SUPABASE_SERVICE_KEY,
            "apikey": SUPABASE_SERVICE_KEY,
            "Content-Type": "application/json"}


def qb_ler():
    r = requests.get(f"{SUPABASE_URL}/rest/v1/qb_tokens",
                     params={"id": "eq.1", "select": "*"},
                     headers=qb_sb_headers(), timeout=30)
    linhas = r.json() if r.ok else []
    return linhas[0] if linhas else None


def qb_gravar(realm, refresh):
    r = requests.post(f"{SUPABASE_URL}/rest/v1/qb_tokens",
                      headers={**qb_sb_headers(),
                               "Prefer": "resolution=merge-duplicates"},
                      json={"id": 1, "realm_id": realm,
                            "refresh_token": refresh},
                      timeout=30)
    if not r.ok:
        log.error("Falha ao salvar o token do QuickBooks: %s %s",
                  r.status_code, (r.text or "")[:300])
        raise HTTPException(500,
            "A autorização funcionou, mas não consegui salvar o token. "
            "A tabela qb_tokens existe no Supabase? Detalhe: "
            + (r.text or "")[:200])


RECONECTAR = ("A autorização do QuickBooks expirou ou foi revogada. "
              "Abra /qb/connect para reconectar.")


def qb_token(codigo=None, refresh=None):
    """Troca código ou refresh por um token. Uma nova tentativa em falha passageira."""
    _, url = qb_endpoints()
    dados = ({"grant_type": "authorization_code", "code": codigo,
              "redirect_uri": QB_REDIRECT} if codigo else
             {"grant_type": "refresh_token", "refresh_token": refresh})
    ultimo = ""
    for tentativa in (1, 2):
        try:
            r = requests.post(url, data=dados, auth=(QB_ID, QB_SECRET),
                              headers={"Accept": "application/json"}, timeout=60)
        except requests.RequestException as e:
            ultimo = str(e)
            time.sleep(1.5)
            continue
        if r.ok:
            return r.json()
        corpo = (r.text or "")[:300]
        log.warning("QBO token erro %s | intuit_tid=%s | %s", r.status_code,
                    r.headers.get("intuit_tid", ""), corpo)
        # invalid_grant = refresh expirado ou acesso revogado: não adianta repetir
        if r.status_code == 400 and "invalid_grant" in corpo:
            _qb_cache.update({"access": None, "exp": 0})
            raise HTTPException(428, RECONECTAR)
        if r.status_code in (401, 403):
            raise HTTPException(428, RECONECTAR)
        ultimo = corpo
        if r.status_code < 500:
            break
        time.sleep(1.5)
    raise HTTPException(502, f"QuickBooks recusou o token: {ultimo}")


def qb_acesso(forcar=False):
    """Devolve (access_token, realm_id).

    O token de acesso dura 1 hora: guardamos em memória e só renovamos
    quando falta pouco. O refresh gira sozinho e é sempre regravado.
    """
    agora = time.time()
    if (not forcar and _qb_cache["access"] and _qb_cache["realm"]
            and agora < _qb_cache["exp"]):
        return _qb_cache["access"], _qb_cache["realm"]
    linha = qb_ler()
    if not linha or not linha.get("refresh_token"):
        raise HTTPException(428, "QuickBooks não está conectado. "
                                 "Abra /qb/connect uma vez para autorizar.")
    j = qb_token(refresh=linha["refresh_token"])
    novo = j.get("refresh_token")
    if novo and novo != linha["refresh_token"]:
        qb_gravar(linha["realm_id"], novo)
    dur = int(j.get("expires_in") or 3600)
    _qb_cache.update({"access": j["access_token"], "realm": linha["realm_id"],
                      "exp": agora + max(60, dur - 300)})
    return _qb_cache["access"], _qb_cache["realm"]


# ---------------------------------------------------------- chamadas
def qb_get(caminho, params=None):
    """Chama a API. Se o token tiver expirado, renova e tenta de novo uma vez."""
    p = dict(params or {})
    p["minorversion"] = QB_MINOR
    for tentativa in (1, 2):
        tok, realm = qb_acesso(forcar=(tentativa == 2))
        r = requests.get(f"{QB_API}/{realm}/{caminho}",
                         headers={"Authorization": "Bearer " + tok,
                                  "Accept": "application/json"},
                         params=p, timeout=90)
        tid = r.headers.get("intuit_tid", "")
        if r.ok:
            return r.json()
        log.warning("QBO erro %s em %s | intuit_tid=%s | %s",
                    r.status_code, caminho, tid, (r.text or "")[:400])
        if r.status_code == 401 and tentativa == 1:
            continue          # token venceu: renova e repete
        if r.status_code == 429:
            log.warning("QBO limite de chamadas atingido | intuit_tid=%s", tid)
            time.sleep(2)
            continue          # limite de chamadas: espera e repete
        raise HTTPException(502, f"QuickBooks ({r.status_code}) "
                                 f"[intuit_tid {tid or 'n/d'}]: {r.text[:200]}")
    raise HTTPException(502, "QuickBooks não respondeu.")


def qb_query(sql):
    return qb_get("query", {"query": sql}).get("QueryResponse", {})


def _esc_sql(v):
    """Escapa para a linguagem de consulta do QuickBooks (barra invertida)."""
    return (str(v or "").replace("\\", "\\\\").replace("'", "\\'"))


def _chave_busca(v):
    """Termo seguro para o LIKE: corta no apóstrofo.

    Endereços como "97 Lunn's Way" viram "97 Lunn", que já é específico o
    bastante e não depende de escape nenhum.
    """
    t = _re.sub(r"\s+", " ", str(v or "")).strip()
    if "'" in t or "\u2019" in t:
        corte = _re.split(r"['\u2019]", t)[0].strip()
        if len(corte) >= 5:
            return corte
    return t


# ---------------------------------------------------------- busca do projeto
def qb_por_estimate(num):
    """Tenta achar pelo número da estimate.

    O QuickBooks NÃO permite consultar pelo memo (PrivateNote não é
    pesquisável), então o caminho é a entidade Estimate pelo número do
    documento — só funciona se a estimate também existir dentro do QBO.
    Quando não existe, devolve None e quem manda é o endereço.
    """
    try:
        q = ("select Id, DocNumber, CustomerRef from Estimate "
             "where DocNumber = '" + _esc_sql(num) + "' maxresults 5")
        est = (qb_query(q).get("Estimate") or [])
        for e in est:
            ref = e.get("CustomerRef") or {}
            if ref.get("value"):
                return {"id": ref["value"], "nome": ref.get("name", ""),
                        "via": "estimate"}
    except HTTPException:
        pass
    return None


_inv_cache = {"quando": 0, "dados": None}


def qb_faturas_cache():
    """Guarda as faturas recentes por 10 min — usada para casar pelo memo."""
    agora = time.time()
    if _inv_cache["dados"] and (agora - _inv_cache["quando"]) < 600:
        return _inv_cache["dados"]
    d = qb_faturas_recentes()
    _inv_cache.update({"quando": agora, "dados": d})
    return d


def qb_por_memo(num):
    """Acha o projeto pelo número da estimate escrito no memo da fatura.

    O QuickBooks não deixa consultar o memo, então varremos as faturas que já
    baixamos. É a ligação mais confiável: independe do endereço, que muda
    quando a obra não é na casa onde o cliente mora.
    """
    n = str(num or "").strip()
    if not n:
        return None
    padrao = _re.compile(r"estimate\s*#?\s*" + _re.escape(n) + r"(?!\d)", _re.I)
    for i in qb_faturas_cache():
        if padrao.search(i.get("PrivateNote") or ""):
            ref = i.get("CustomerRef") or {}
            if ref.get("value"):
                return {"id": ref["value"], "nome": ref.get("name", ""),
                        "via": "estimate"}
    return None


# ---------------------------------------------------------- mão de obra
_tempo_cache = {"quando": 0, "dados": None}


def qb_tempo_recente(meses=24):
    """Todos os apontamentos de hora dos últimos meses."""
    from datetime import timedelta
    corte = (_date.today() - timedelta(days=int(meses * 30.5))).isoformat()
    todas, pos = [], 1
    for _ in range(6):
        q = ("select * from TimeActivity where TxnDate >= '" + corte + "' "
             "startposition " + str(pos) + " maxresults 1000")
        lote = qb_query(q).get("TimeActivity") or []
        todas += lote
        if len(lote) < 1000:
            break
        pos += 1000
    return todas


def qb_tempo_cache():
    agora = time.time()
    if _tempo_cache["dados"] and (agora - _tempo_cache["quando"]) < 600:
        return _tempo_cache["dados"]
    try:
        d = qb_tempo_recente()
    except HTTPException as e:
        log.warning("Não consegui ler as horas: %s", e.detail)
        d = []
    _tempo_cache.update({"quando": agora, "dados": d})
    return d


def _pessoa_de(t):
    for k in ("EmployeeRef", "VendorRef"):
        r = t.get(k) or {}
        if r.get("name"):
            return r["name"]
    return (t.get("NameOf") or "").strip() or "—"


_auto_cache = {"quando": 0, "dados": None}


def qb_rates_auto():
    """Descobre o valor/hora sozinho, sem ninguém digitar.

    Procura em três lugares, nesta ordem:
      1. CostRate gravado no próprio apontamento de hora
      2. HourlyRate do apontamento
      3. BillRate do cadastro do funcionário ou do prestador
    """
    agora = time.time()
    if _auto_cache["dados"] and (agora - _auto_cache["quando"]) < 600:
        return _auto_cache["dados"]

    achado = {}

    def guarda(nome, valor, origem):
        v = float(valor or 0)
        if nome and v > 0 and nome not in achado:
            achado[nome] = {"rate": v, "origem": origem}

    for t in qb_tempo_cache():
        n = _pessoa_de(t)
        guarda(n, t.get("CostRate"), "apontamento (CostRate)")
    for t in qb_tempo_cache():
        n = _pessoa_de(t)
        guarda(n, t.get("HourlyRate"), "apontamento (HourlyRate)")

    for entidade in ("Employee", "Vendor"):
        try:
            reg = qb_query("select * from " + entidade
                           + " where Active = true maxresults 500")
            for e in (reg.get(entidade) or []):
                nome = (e.get("DisplayName")
                        or " ".join(x for x in [e.get("GivenName"),
                                                e.get("FamilyName")] if x))
                guarda(nome, e.get("BillRate"), entidade + ".BillRate")
        except HTTPException as ex:
            log.warning("Não li %s: %s", entidade, ex.detail)

    _auto_cache.update({"quando": agora, "dados": achado})
    return achado


def qb_ler_rates():
    r = requests.get(f"{SUPABASE_URL}/rest/v1/qb_rates",
                     params={"select": "nome,rate"},
                     headers=qb_sb_headers(), timeout=30)
    linhas = r.json() if r.ok else []
    return {x["nome"]: float(x.get("rate") or 0) for x in linhas}


def qb_horas_projeto(cid):
    """Horas e custo de mão de obra de um projeto.

    O valor/hora vem da tabela qb_rates; se a pessoa não estiver lá, usa o
    que estiver gravado no próprio apontamento (CostRate/HourlyRate).
    """
    rates = qb_ler_rates()
    auto = qb_rates_auto()
    por_pessoa, horas_tot, custo_tot = {}, 0.0, 0.0
    for t in qb_tempo_cache():
        ref = (t.get("CustomerRef") or {}).get("value")
        if str(ref or "") != str(cid):
            continue
        h = float(t.get("Hours") or 0) + (float(t.get("Minutes") or 0) / 60.0)
        if h <= 0:
            continue
        nome = _pessoa_de(t)
        taxa = rates.get(nome)
        if not taxa:
            a = auto.get(nome)
            taxa = a["rate"] if a else 0.0
        d = por_pessoa.setdefault(nome, {"nome": nome, "horas": 0.0,
                                         "taxa": taxa, "custo": 0.0})
        d["horas"] += h
        d["custo"] += h * taxa
        horas_tot += h
        custo_tot += h * taxa
    lista = sorted(por_pessoa.values(), key=lambda x: -x["custo"])
    for d in lista:
        d["horas"] = round(d["horas"], 2)
        d["custo"] = round(d["custo"], 2)
    return {"horas": round(horas_tot, 2), "custo": round(custo_tot, 2),
            "pessoas": lista,
            "sem_taxa": [d["nome"] for d in lista if not d["taxa"]]}


@app.get("/qb/taxas")
def qb_taxas(authorization: str = Header(default="")):
    """Mostra o valor/hora encontrado de cada pessoa e a origem."""
    quem_e(authorization)
    manual = qb_ler_rates()
    auto = qb_rates_auto()
    horas = {}
    campos = set()
    for t in qb_tempo_cache():
        h = float(t.get("Hours") or 0) + (float(t.get("Minutes") or 0) / 60.0)
        n = _pessoa_de(t)
        if h > 0:
            horas[n] = horas.get(n, 0.0) + h
        campos.update(k for k in t.keys() if "ate" in k or "Rate" in k)
    saida = []
    for n, h in sorted(horas.items(), key=lambda x: -x[1]):
        a = auto.get(n)
        saida.append({"nome": n, "horas": round(h, 1),
                      "rate": manual.get(n) or (a["rate"] if a else 0),
                      "origem": ("cadastro manual" if manual.get(n)
                                 else (a["origem"] if a else "NÃO ENCONTRADO"))})
    return {"ok": True, "pessoas": saida,
            "campos_no_apontamento": sorted(campos)}


@app.get("/qb/pessoas")
def qb_pessoas(authorization: str = Header(default="")):
    """Quem apontou hora e quanto está cadastrado por hora."""
    quem_e(authorization)
    rates = qb_ler_rates()
    horas = {}
    for t in qb_tempo_cache():
        h = float(t.get("Hours") or 0) + (float(t.get("Minutes") or 0) / 60.0)
        if h <= 0:
            continue
        n = _pessoa_de(t)
        horas[n] = horas.get(n, 0.0) + h
    auto = qb_rates_auto()
    lista = []
    for n, v in sorted(horas.items(), key=lambda x: -x[1]):
        a = auto.get(n)
        lista.append({"nome": n, "horas": round(v, 1),
                      "rate": rates.get(n) or (a["rate"] if a else 0),
                      "origem": ("manual" if rates.get(n)
                                 else (a["origem"] if a else ""))})
    return {"ok": True, "pessoas": lista}


class RatesReq(BaseModel):
    rates: dict[str, float] = {}


@app.post("/qb/pessoas")
def qb_pessoas_salvar(req: RatesReq, authorization: str = Header(default="")):
    _, papel, _ = quem_e(authorization)
    if papel != "admin":
        raise HTTPException(403, "Só admin altera o custo por hora.")
    linhas = [{"nome": n, "rate": float(v or 0)} for n, v in req.rates.items()]
    if linhas:
        r = requests.post(f"{SUPABASE_URL}/rest/v1/qb_rates",
                          headers={**qb_sb_headers(),
                                   "Prefer": "resolution=merge-duplicates"},
                          json=linhas, timeout=30)
        if not r.ok:
            raise HTTPException(500, "Não consegui salvar: " + r.text[:200])
    return {"ok": True, "salvos": len(linhas)}


@app.get("/qb/buscar")
def qb_buscar(q: str = "", authorization: str = Header(default="")):
    """Lista projetos do QuickBooks por parte do nome — para ligar na mão."""
    quem_e(authorization)
    termo = _chave_busca(q)
    if len(termo) < 3:
        return {"ok": True, "projetos": []}
    sql = ("select Id, DisplayName from Customer where Active = true "
           "and DisplayName like '%" + _esc_sql(termo) + "%' maxresults 20")
    try:
        lista = (qb_query(sql).get("Customer") or [])
    except HTTPException as e:
        return {"ok": False, "motivo": str(e.detail)}
    return {"ok": True, "projetos": [{"id": c["Id"],
                                      "nome": c.get("DisplayName", "")}
                                     for c in lista]}


_ABREV = {
    "AVE": "AVENUE", "AV": "AVENUE", "ST": "STREET", "STR": "STREET",
    "RD": "ROAD", "DR": "DRIVE", "LN": "LANE", "CT": "COURT",
    "CIR": "CIRCLE", "PL": "PLACE", "TER": "TERRACE", "TERR": "TERRACE",
    "HWY": "HIGHWAY", "BLVD": "BOULEVARD", "PKWY": "PARKWAY",
    "SQ": "SQUARE", "TRL": "TRAIL", "PSGE": "PASSAGE", "PSG": "PASSAGE",
    "XING": "CROSSING", "RT": "ROUTE", "EXT": "EXTENSION",
}


def _expande(t):
    """Mesma normalização do app: abreviação vira palavra inteira.

    Sem isto, "11 Bosuns Psge" no card nunca encontra "11 Bosuns Passage"
    no QuickBooks — era essa a causa de projeto "não encontrado" mesmo
    existindo.
    """
    p = _re.sub(r"[.,;:]+", " ", str(t or "").upper())
    p = _re.sub(r"\s+", " ", p).strip().split(" ")
    return [_ABREV.get(w, w) for w in p if w]


def _numero_e_rua(rua):
    """Devolve (número da casa, palavras da rua) já expandidas.

    O número nem sempre é a primeira coisa: "MCINTYRE 21 Hidden Valley"
    e "STAIRS - 6539 2 Daniel Lane" existem na agenda. Vale o primeiro
    número seguido de uma palavra.
    """
    p = _expande(rua)
    for i, w in enumerate(p):
        if not _re.fullmatch(r"\d+[A-Z]?", w):
            continue
        resto = p[i + 1:]
        if resto and _re.match(r"^[A-Z]", resto[0]):
            return w, resto
    return "", []


def qb_por_endereco(rua, cidade=""):
    """Acha o projeto pelo endereço que está dentro do nome do projeto.

    Estratégia: procurar no QuickBooks só pela parte que não varia —
    número + primeira palavra da rua — e conferir o resto aqui, onde dá
    para expandir abreviações. Procurar a rua inteira falhava sempre que
    os dois lados escreviam "Passage" e "Psge" de formas diferentes.
    """
    num, palavras = _numero_e_rua(rua)
    if not num or not palavras:
        return None

    # A busca precisa de algo específico. Uma palavra basta na maioria;
    # se for curtinha ("A", "Old"), usa duas.
    quantas = 1 if len(palavras[0]) >= 4 else 2
    alvo = _chave_busca(num + " " + " ".join(palavras[:quantas]))
    if len(alvo) < 4:
        return None

    vistos, lista = set(), []
    for termo in (alvo, num + " " + palavras[0]):
        termo = _chave_busca(termo)
        if termo in vistos or len(termo) < 4:
            continue
        vistos.add(termo)
        q = ("select Id, DisplayName from Customer where Active = true "
             "and DisplayName like '%" + _esc_sql(termo) + "%' maxresults 30")
        try:
            lista += (qb_query(q).get("Customer") or [])
        except HTTPException:
            pass
        if lista:
            break
    if not lista:
        return None

    def pontos(x):
        """Quanto este candidato combina com o endereço do card."""
        nome = _expande(x.get("DisplayName", ""))
        nnum, nrua = _numero_e_rua(" ".join(nome))
        # número da casa diferente elimina: "11 Bosuns" x "111 Bosuns"
        if nnum and nnum != num:
            return -1
        if num not in nome:
            return -1
        p = 0
        for w in palavras[:3]:
            if w in nome:
                p += 2
        if cidade:
            for w in _expande(cidade):
                if w in nome:
                    p += 1
        return p

    marcados = [(pontos(x), i, x) for i, x in enumerate(lista)]
    marcados = [m for m in marcados if m[0] >= 2]      # ao menos a rua bate
    if not marcados:
        return None
    marcados.sort(key=lambda m: (-m[0], m[1]))
    x = marcados[0][2]
    return {"id": x["Id"], "nome": x.get("DisplayName", ""), "via": "endereco",
            "outros": [y[2].get("DisplayName", "") for y in marcados[1:5]]}


# ---------------------------------------------------------- números
def qb_faturas(cid):
    q = ("select Id, DocNumber, TotalAmt, Balance, TxnDate, DueDate, PrivateNote "
         "from Invoice where CustomerRef = '" + _esc_sql(cid) + "' maxresults 200")
    inv = (qb_query(q).get("Invoice") or [])
    faturado = sum(float(i.get("TotalAmt") or 0) for i in inv)
    saldo = sum(float(i.get("Balance") or 0) for i in inv)
    itens = [{"id": i.get("Id"),
              "num": i.get("DocNumber"), "valor": float(i.get("TotalAmt") or 0),
              "saldo": float(i.get("Balance") or 0), "data": i.get("TxnDate"),
              "vence": i.get("DueDate"), "memo": i.get("PrivateNote", "")}
             for i in inv]
    itens.sort(key=lambda x: x.get("data") or "")
    return {"faturado": faturado, "saldo": saldo, "recebido": faturado - saldo,
            "faturas": itens}


def _pl_valor(rows, alvo):
    """Varre o relatório e devolve o total do grupo pedido."""
    for row in (rows or []):
        if row.get("group") == alvo:
            col = ((row.get("Summary") or {}).get("ColData") or [])
            if len(col) > 1:
                try:
                    return float(str(col[1].get("value") or 0).replace(",", ""))
                except ValueError:
                    return 0.0
        filhos = (row.get("Rows") or {}).get("Row")
        if filhos:
            v = _pl_valor(filhos, alvo)
            if v is not None:
                return v
    return None


def qb_custos(cid):
    """Lucro do projeto pelo relatório de P&L filtrado por aquele projeto."""
    rel = qb_get("reports/ProfitAndLoss",
                 {"customer": cid, "start_date": "2015-01-01",
                  "end_date": _date.today().isoformat(),
                  "accounting_method": "Accrual"})
    linhas = (rel.get("Rows") or {}).get("Row") or []
    receita = _pl_valor(linhas, "Income") or 0.0
    cogs = _pl_valor(linhas, "COGS") or 0.0
    desp = _pl_valor(linhas, "Expenses") or 0.0
    liquido = _pl_valor(linhas, "NetIncome")
    custos = cogs + desp
    if liquido is None:
        liquido = receita - custos
    return {"receita": receita, "custos": custos, "cogs": cogs,
            "despesas": desp, "lucro": liquido}


# ---------------------------------------------------------- rotas
@app.get("/qb/connect")
def qb_connect():
    if not QB_ID or not SERVICE_URL:
        raise HTTPException(500, "Faltam QB_CLIENT_ID / SERVICE_URL no Render.")
    estado = _secrets.token_urlsafe(24)
    _qb_estado[estado] = time.time()
    # limpa estados velhos (mais de 15 min)
    for k in [k for k, v in _qb_estado.items() if time.time() - v > 900]:
        _qb_estado.pop(k, None)
    auth_url, _ = qb_endpoints()
    url = (auth_url + "?client_id=" + QB_ID +
           "&response_type=code&scope=com.intuit.quickbooks.accounting" +
           "&redirect_uri=" + requests.utils.quote(QB_REDIRECT, safe="") +
           "&state=" + estado)
    return RedirectResponse(url)


@app.get("/qb/callback")
def qb_callback(code: str = "", state: str = "", realmId: str = ""):
    if not code or not realmId:
        return HTMLResponse("<h3>Autorização cancelada.</h3>", status_code=400)
    # CSRF: o state tem que ser um que nós mesmos criamos
    if not state or state not in _qb_estado:
        return HTMLResponse(
            "<div style='font-family:system-ui;padding:40px;text-align:center'>"
            "<h2>Pedido inválido</h2><p>O código de segurança (state) não confere. "
            "Comece de novo em <a href='/qb/connect'>/qb/connect</a>.</p></div>",
            status_code=400)
    _qb_estado.pop(state, None)
    j = qb_token(codigo=code)
    qb_gravar(realmId, j.get("refresh_token"))
    return HTMLResponse(
        "<div style='font-family:system-ui;padding:40px;text-align:center'>"
        "<h2>QuickBooks conectado</h2>"
        "<p>Empresa " + realmId + ". Pode fechar esta aba.</p></div>")


@app.get("/qb/disconnect")
def qb_disconnect():
    """A Intuit exige um endereço de desconexão. Apaga o token guardado."""
    try:
        requests.delete(f"{SUPABASE_URL}/rest/v1/qb_tokens",
                        params={"id": "eq.1"}, headers=qb_sb_headers(),
                        timeout=30)
    except Exception:
        pass
    _qb_cache.update({"access": None, "exp": 0, "realm": None})
    return HTMLResponse(
        "<div style='font-family:system-ui;padding:40px;text-align:center'>"
        "<h2>QuickBooks disconnected</h2>"
        "<p>PLJ Schedule no longer has access to your QuickBooks data.</p>"
        "<p><a href='/qb/connect'>Connect again</a></p></div>")


@app.get("/qb/status")
def qb_status():
    linha = qb_ler()
    if not linha:
        return {"conectado": False,
                "motivo": "Nenhum token salvo. Rode o qb_tokens.sql no Supabase "
                          "e depois abra /qb/connect de novo."}
    try:
        tok, realm = qb_acesso()
        info = qb_get("companyinfo/" + realm).get("CompanyInfo", {})
        return {"conectado": True, "realm": realm,
                "empresa": info.get("CompanyName", "")}
    except HTTPException as e:
        return {"conectado": False, "erro": e.detail}


CAMPOS_INV = ("Id, DocNumber, CustomerRef, TotalAmt, Balance, TxnDate, "
              "DueDate, PrivateNote")


def qb_faturas_recentes(meses=18):
    """Faturas dos últimos meses + qualquer uma antiga ainda em aberto.

    Traz também as já pagas, porque a planilha precisa mostrar 'Pago'.
    Pagina de mil em mil, que é o teto do QuickBooks por consulta.
    """
    from datetime import timedelta
    corte = (_date.today() - timedelta(days=int(meses * 30.5))).isoformat()
    vistas, todas = set(), []

    def junta(lista):
        for i in lista:
            k = i.get("Id")
            if k and k not in vistas:
                vistas.add(k)
                todas.append(i)

    pos = 1
    for _ in range(6):
        q = (f"select {CAMPOS_INV} from Invoice where TxnDate >= '{corte}' "
             f"order by TxnDate startposition {pos} maxresults 1000")
        lote = qb_query(q).get("Invoice") or []
        junta(lote)
        if len(lote) < 1000:
            break
        pos += 1000

    # em aberto mais antigas que o corte
    q2 = (f"select {CAMPOS_INV} from Invoice where Balance > '0' "
          f"and TxnDate < '{corte}' maxresults 1000")
    try:
        junta(qb_query(q2).get("Invoice") or [])
    except HTTPException:
        pass
    return todas


def _pl_por_coluna(linhas, alvo, qtd):
    """Acha a linha de um grupo e devolve os valores de cada coluna."""
    for row in (linhas or []):
        if row.get("group") == alvo:
            col = ((row.get("Summary") or {}).get("ColData") or [])
            out = []
            for i in range(qtd):
                v = col[i].get("value") if i < len(col) else ""
                try:
                    out.append(float(str(v or 0).replace(",", "")))
                except ValueError:
                    out.append(0.0)
            return out
        filhos = (row.get("Rows") or {}).get("Row")
        if filhos:
            v = _pl_por_coluna(filhos, alvo, qtd)
            if v is not None:
                return v
    return None


def qb_margens():
    """Receita, custo e margem de TODOS os projetos numa única chamada.

    Usa o P&L quebrado por cliente (summarize_column_by=Customers), em vez de
    um relatório por obra — seriam dezenas de chamadas.
    """
    rel = qb_get("reports/ProfitAndLoss",
                 {"summarize_column_by": "Customers",
                  "start_date": "2015-01-01",
                  "end_date": _date.today().isoformat(),
                  "accounting_method": "Accrual"})
    cols = (rel.get("Columns") or {}).get("Column") or []
    titulos = [c.get("ColTitle") or "" for c in cols]
    n = len(titulos)
    linhas = (rel.get("Rows") or {}).get("Row") or []
    receita = _pl_por_coluna(linhas, "Income", n) or [0.0] * n
    cogs = _pl_por_coluna(linhas, "COGS", n) or [0.0] * n
    desp = _pl_por_coluna(linhas, "Expenses", n) or [0.0] * n
    liq = _pl_por_coluna(linhas, "NetIncome", n)

    saida = {}
    for i, t in enumerate(titulos):
        if not t or t.lower() in ("total", ""):
            continue
        r = receita[i] if i < len(receita) else 0.0
        c = (cogs[i] if i < len(cogs) else 0.0) + (desp[i] if i < len(desp) else 0.0)
        l = (liq[i] if (liq and i < len(liq)) else (r - c))
        saida[t] = {"receita": round(r, 2), "custos": round(c, 2),
                    "lucro": round(l, 2),
                    "margem": (round((l / r) * 100, 1) if r else None),
                    "pct_gasto": (round((c / r) * 100, 1) if r else None)}
    return saida


@app.get("/qb/margens")
def qb_margens_rota(authorization: str = Header(default="")):
    quem_e(authorization)
    try:
        rel = qb_get("reports/ProfitAndLoss",
                     {"summarize_column_by": "Customers",
                      "start_date": "2015-01-01",
                      "end_date": _date.today().isoformat(),
                      "accounting_method": "Accrual"})
        cols = (rel.get("Columns") or {}).get("Column") or []
        titulos = [c.get("ColTitle") or "" for c in cols]
        grupos = []

        def varre(rows, nivel=0):
            for r in (rows or []):
                g = r.get("group") or r.get("type") or ""
                if g:
                    grupos.append(g)
                f = (r.get("Rows") or {}).get("Row")
                if f and nivel < 2:
                    varre(f, nivel + 1)
        varre((rel.get("Rows") or {}).get("Row") or [])
        return {"ok": True, "projetos": qb_margens(),
                "debug": {"colunas": titulos[:12], "qtd_colunas": len(titulos),
                          "grupos": sorted(set(grupos))[:14]}}
    except HTTPException as e:
        return {"ok": False, "motivo": str(e.detail)}


@app.get("/qb/abertas")
def qb_abertas(authorization: str = Header(default="")):
    """Todas as faturas em aberto, somadas por projeto. Uma chamada só."""
    quem_e(authorization)          # basta estar cadastrado no schedule
    inv = qb_faturas_recentes()
    por = {}
    for i in inv:
        ref = i.get("CustomerRef") or {}
        nome = ref.get("name") or ""
        cid = ref.get("value") or ""
        if not nome:
            continue
        d = por.setdefault(nome, {"projeto": nome, "id": cid,
                                  "saldo": 0.0, "faturado": 0.0, "qtd": 0,
                                  "vence": None, "faturas": []})
        saldo = float(i.get("Balance") or 0)
        d["saldo"] += saldo
        d["faturado"] += float(i.get("TotalAmt") or 0)
        d["qtd"] += 1
        v = i.get("DueDate")
        if v and (d["vence"] is None or v < d["vence"]):
            d["vence"] = v
        d["faturas"].append({"num": i.get("DocNumber"),
                             "valor": round(float(i.get("TotalAmt") or 0), 2),
                             "saldo": round(saldo, 2),
                             "data": i.get("TxnDate"), "vence": v,
                             "memo": i.get("PrivateNote", "")})
    lista = [d for d in por.values() if d["saldo"] > 0.004]
    lista.sort(key=lambda x: -x["saldo"])
    for d in lista:
        d["saldo"] = round(d["saldo"], 2)
        d["faturado"] = round(d["faturado"], 2)
        d["faturas"].sort(key=lambda f: f.get("data") or "")
    return {"ok": True, "total": round(sum(d["saldo"] for d in lista), 2),
            "projetos": lista}


class ProjReq(BaseModel):
    num: str = ""
    rua: str = ""
    cidade: str = ""
    qbid: str = ""


@app.post("/qb/projeto")
def qb_projeto(req: ProjReq, authorization: str = Header(default="")):
    quem_e(authorization)          # basta estar cadastrado no schedule
    proj = None
    # 1) ligação feita à mão manda em tudo
    if req.qbid:
        nome = ""
        try:
            c = (qb_query("select Id, DisplayName from Customer where Id = '"
                          + _esc_sql(req.qbid) + "'").get("Customer") or [])
            nome = c[0].get("DisplayName", "") if c else ""
        except HTTPException:
            pass
        proj = {"id": req.qbid, "nome": nome or ("Projeto " + req.qbid),
                "via": "manual"}
    # 2) número da estimate (independe do endereço)
    if not proj and req.num:
        for tentativa in (qb_por_memo, qb_por_estimate):
            try:
                proj = tentativa(req.num)
            except HTTPException as e:
                log.warning("Busca por estimate falhou (%s): %s",
                            req.num, e.detail)
            if proj:
                break
    # 3) endereço, como última opção
    if not proj and req.rua:
        try:
            proj = qb_por_endereco(req.rua, req.cidade)
        except HTTPException as e:
            log.warning("Busca por endereço falhou (%s): %s", req.rua, e.detail)
    if not proj:
        return {"ok": False,
                "motivo": ("Projeto não encontrado no QuickBooks. "
                           "Preencha o número da estimate, ou use "
                           "\"Link project\" para escolher na mão.")}

    fat = qb_faturas(proj["id"])
    try:
        pl = qb_custos(proj["id"])
    except HTTPException:
        pl = {"receita": fat["faturado"], "custos": 0.0, "lucro": 0.0}

    confere = None
    if req.num:
        marca = "estimate " + str(req.num).strip().lower()
        bate = [f for f in fat["faturas"]
                if marca in (f.get("memo") or "").lower()]
        confere = bool(bate)

    try:
        mo = qb_horas_projeto(proj["id"])
    except HTTPException:
        mo = {"horas": 0.0, "custo": 0.0, "pessoas": [], "sem_taxa": []}

    custos = pl["custos"] + mo["custo"]
    lucro = pl["lucro"] - mo["custo"]
    base = pl["receita"] or fat["faturado"]
    pct = round((custos / base) * 100, 1) if base else None
    margem = round((lucro / base) * 100, 1) if base else None
    return {"ok": True, "projeto": proj["nome"], "id": proj["id"],
            "via": proj.get("via"), "outros": proj.get("outros", []),
            "confere_estimate": confere,
            "faturado": round(fat["faturado"], 2),
            "recebido": round(fat["recebido"], 2),
            "deve": round(fat["saldo"], 2),
            "custos": round(custos, 2),
            "custos_material": round(pl["custos"], 2),
            "mao_obra": round(mo["custo"], 2),
            "horas": mo["horas"], "pessoas": mo["pessoas"],
            "sem_taxa": mo["sem_taxa"],
            "lucro": round(lucro, 2),
            "pct_gasto": pct, "margem": margem,
            "faturas": fat["faturas"][:12]}


# =====================================================================
# PLANILHA DE COBRANÇA (.xlsx) — no formato que a equipe já usa
# =====================================================================
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VERDE = "C6E0B4"
VERM = "FF5B5B"
AMAR = "FFE699"
CINZA = "D9D9D9"
ESCURO = "404040"
BRANCO = "FFFFFF"

COLUNAS = [("Cliente", 30), ("Estimate", 11), ("Down", 12), ("Middle", 12),
           ("Data Middle", 13), ("Situação", 15), ("Final", 12),
           ("Data Final", 13), ("Situação", 15), ("Status", 16),
           ("Total em aberto", 15), ("Observação", 52)]

BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)


class LinhaPlan(BaseModel):
    pm: str = ""
    cliente: str = ""
    estimate: str = ""
    down: float = 0
    middle: float = 0
    data_middle: str = ""
    sit_middle: str = ""
    final: float = 0
    data_final: str = ""
    sit_final: str = ""
    status: str = ""
    aberto: float = 0
    obs: str = ""


class PlanReq(BaseModel):
    titulo: str = "Cobranças"
    linhas: list[LinhaPlan] = []


def _cor_situacao(txt):
    t = (txt or "").lower()
    if "pago" in t:
        return VERDE
    if "atras" in t:
        return VERM
    if "prazo" in t:
        return VERDE
    if "sem data" in t:
        return AMAR
    return None


def monta_planilha(req):
    """Monta o arquivo e devolve os bytes. Separado para poder testar."""
    wb = Workbook()
    wb.remove(wb.active)

    por_pm = {}
    for l in req.linhas:
        por_pm.setdefault(l.pm or "Sem PM", []).append(l)

    ordem = sorted(por_pm.keys())
    if len(ordem) > 1:
        ordem = ["TODOS"] + ordem
        por_pm["TODOS"] = list(req.linhas)

    for pm in ordem:
        linhas = por_pm[pm]
        aba = wb.create_sheet((pm or "Sem PM")[:28] or "Sem PM")

        # faixa com o nome do PM
        aba.merge_cells(start_row=1, start_column=1,
                        end_row=1, end_column=len(COLUNAS))
        c = aba.cell(row=1, column=1, value=pm.upper())
        c.font = Font(bold=True, size=13, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=ESCURO)
        c.alignment = Alignment(horizontal="center", vertical="center")
        aba.row_dimensions[1].height = 24

        # cabeçalho
        for i, (nome, larg) in enumerate(COLUNAS, start=1):
            h = aba.cell(row=2, column=i, value=nome)
            h.font = Font(bold=True, size=10)
            h.fill = PatternFill("solid", fgColor=CINZA)
            h.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            h.border = BORDA
            aba.column_dimensions[get_column_letter(i)].width = larg
        aba.row_dimensions[2].height = 28

        linhas.sort(key=lambda x: ((x.cliente or "").upper(), x.estimate or ""))
        r = 3
        tot_aberto = 0.0
        for l in linhas:
            atrasado = ("atras" in (l.sit_middle or "").lower()
                        or "atras" in (l.sit_final or "").lower())
            valores = [l.cliente, l.estimate, l.down or None, l.middle or None,
                       l.data_middle, l.sit_middle, l.final or None,
                       l.data_final, l.sit_final, l.status,
                       l.aberto or None, l.obs]
            for i, v in enumerate(valores, start=1):
                cel = aba.cell(row=r, column=i, value=v)
                cel.border = BORDA
                cel.font = Font(size=10, bold=(i == 1))
                if i in (3, 4, 7, 11):
                    cel.number_format = '"$"#,##0.00'
                    cel.alignment = Alignment(horizontal="right")
                elif i in (2, 5, 6, 8, 9, 10):
                    cel.alignment = Alignment(horizontal="center")
                else:
                    cel.alignment = Alignment(vertical="center", wrap_text=True)
                if i in (6, 9):
                    cor = _cor_situacao(v)
                    if cor:
                        cel.fill = PatternFill("solid", fgColor=cor)
                        cel.font = Font(size=10, bold=True)
                if i == 10:
                    st = (v or "").lower()
                    if "complet" in st:
                        cel.fill = PatternFill("solid", fgColor=VERDE)
                    elif "hold" in st or "pausad" in st:
                        cel.fill = PatternFill("solid", fgColor=AMAR)
                if i == 12 and atrasado and v:
                    cel.fill = PatternFill("solid", fgColor=VERM)
                    cel.font = Font(size=10, bold=True)
            tot_aberto += float(l.aberto or 0)
            r += 1

        # total
        t = aba.cell(row=r, column=10, value="TOTAL EM ABERTO")
        t.font = Font(bold=True, size=10)
        t.alignment = Alignment(horizontal="right")
        tv = aba.cell(row=r, column=11, value=tot_aberto)
        tv.font = Font(bold=True, size=11)
        tv.number_format = '"$"#,##0.00'
        tv.fill = PatternFill("solid", fgColor=CINZA)
        tv.border = BORDA

        aba.freeze_panes = "A3"
        aba.auto_filter.ref = f"A2:{get_column_letter(len(COLUNAS))}{max(2, r - 1)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.post("/planilha")
def planilha(req: PlanReq, authorization: str = Header(default="")):
    quem_e(authorization)
    dados = monta_planilha(req)
    nome = ("cobrancas_" + _re.sub(r"[^A-Za-z0-9]+", "_", req.titulo or "plj")
            + "_" + _date.today().isoformat() + ".xlsx")
    return StreamingResponse(
        io.BytesIO(dados),
        media_type=("application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"),
        headers={"Content-Disposition": f'attachment; filename="{nome}"'})


# =====================================================================
# CRIAR E ENVIAR FATURA — o app passa a ESCREVER no QuickBooks.
#
# Até aqui tudo era leitura. Daqui para baixo o serviço cria documento
# contábil de verdade, então cada rota confere o papel de quem chamou:
# viewer não passa. O envio é uma chamada separada de propósito — criar
# e mandar para o cliente são duas decisões diferentes.
#
# O escopo com.intuit.quickbooks.accounting, que já está autorizado,
# cobre escrita. Não precisa reconectar o QuickBooks.
# =====================================================================


def qb_post(caminho, corpo=None, params=None, octeto=False):
    """POST na API. Renova o token e repete uma vez se ele tiver vencido."""
    p = dict(params or {})
    p["minorversion"] = QB_MINOR
    for tentativa in (1, 2):
        tok, realm = qb_acesso(forcar=(tentativa == 2))
        cabec = {"Authorization": "Bearer " + tok, "Accept": "application/json"}
        if octeto:
            # a rota de envio da Intuit exige corpo vazio e este content-type
            cabec["Content-Type"] = "application/octet-stream"
            r = requests.post(f"{QB_API}/{realm}/{caminho}", headers=cabec,
                              params=p, data=b"", timeout=90)
        else:
            cabec["Content-Type"] = "application/json"
            r = requests.post(f"{QB_API}/{realm}/{caminho}", headers=cabec,
                              params=p, json=(corpo or {}), timeout=90)
        tid = r.headers.get("intuit_tid", "")
        if r.ok:
            return r.json()
        log.warning("QBO POST erro %s em %s | intuit_tid=%s | %s",
                    r.status_code, caminho, tid, (r.text or "")[:400])
        if r.status_code == 401 and tentativa == 1:
            continue
        detalhe = (r.text or "")[:300]
        raise HTTPException(502, f"QuickBooks recusou ({r.status_code}) "
                                 f"[intuit_tid {tid or 'n/d'}]: {detalhe}")
    raise HTTPException(502, "QuickBooks não respondeu.")


def _pode_faturar(authorization):
    _, papel, nome = quem_e(authorization)
    if papel not in ("admin", "pm"):
        raise HTTPException(403, "Só admin ou project manager pode faturar.")
    return papel, nome


_rec_cache = {}       # usado aqui e pela seção do recebido, mais abaixo


def _limpa_caches_fatura():
    """A fatura nova precisa aparecer na hora, não daqui a 10 minutos."""
    _inv_cache.update({"quando": 0, "dados": None})
    _rec_cache.clear()


@app.get("/qb/itens")
def qb_itens(authorization: str = Header(default="")):
    """Itens de serviço do QuickBooks — é o que o PM escolhe na linha."""
    quem_e(authorization)
    q = ("select Id, Name, Type, UnitPrice, Description from Item "
         "where Active = true maxresults 500")
    lista = (qb_query(q).get("Item") or [])
    saida = []
    for i in lista:
        if (i.get("Type") or "") == "Category":
            continue
        saida.append({"id": i.get("Id"), "nome": i.get("Name", ""),
                      "tipo": i.get("Type", ""),
                      "descricao": i.get("Description", ""),
                      "preco": float(i.get("UnitPrice") or 0)})
    saida.sort(key=lambda x: (x["nome"] or "").lower())
    return {"ok": True, "itens": saida}


@app.get("/qb/cliente")
def qb_cliente(qbid: str = "", authorization: str = Header(default="")):
    """Nome e e-mail do projeto — para pré-preencher o envio."""
    quem_e(authorization)
    if not qbid:
        raise HTTPException(400, "Sem projeto.")
    c = (qb_query("select * from Customer where Id = '" + _esc_sql(qbid) + "'")
         .get("Customer") or [])
    if not c:
        raise HTTPException(404, "Projeto não encontrado.")
    x = c[0]
    email = ((x.get("PrimaryEmailAddr") or {}).get("Address") or "")
    return {"ok": True, "id": x.get("Id"), "nome": x.get("DisplayName", ""),
            "email": email}


def qb_get_pdf(caminho):
    """Baixa um documento em PDF da API (fatura, orçamento)."""
    p = {"minorversion": QB_MINOR}
    for tentativa in (1, 2):
        tok, realm = qb_acesso(forcar=(tentativa == 2))
        r = requests.get(f"{QB_API}/{realm}/{caminho}",
                         headers={"Authorization": "Bearer " + tok,
                                  "Accept": "application/pdf"},
                         params=p, timeout=90)
        tid = r.headers.get("intuit_tid", "")
        if r.ok and r.content[:4] == b"%PDF":
            return r.content
        log.warning("QBO PDF erro %s em %s | intuit_tid=%s | %s",
                    r.status_code, caminho, tid, (r.text or "")[:300])
        if r.status_code == 401 and tentativa == 1:
            continue
        raise HTTPException(502, f"QuickBooks não devolveu o PDF "
                                 f"({r.status_code}) [intuit_tid {tid or 'n/d'}].")
    raise HTTPException(502, "QuickBooks não respondeu.")


class InvLinha(BaseModel):
    item_id: str = ""
    descricao: str = ""
    valor: float = 0


class InvReq(BaseModel):
    qbid: str = ""            # Id do projeto (Customer) no QuickBooks
    num: str = ""             # número da estimate, vai para o memo
    linhas: list[InvLinha] = []
    vence: str = ""           # YYYY-MM-DD
    memo: str = ""


@app.post("/qb/invoice")
def qb_invoice_criar(req: InvReq, authorization: str = Header(default="")):
    """Cria a fatura. NÃO envia nada para o cliente."""
    papel, quem = _pode_faturar(authorization)
    if not req.qbid:
        raise HTTPException(400, "Sem projeto do QuickBooks.")
    linhas = [l for l in req.linhas if float(l.valor or 0) > 0]
    if not linhas:
        raise HTTPException(400, "A fatura precisa de pelo menos uma linha "
                                 "com valor maior que zero.")

    corpo = {"CustomerRef": {"value": str(req.qbid)}, "Line": []}
    for l in linhas:
        valor = round(float(l.valor), 2)
        det = {"Qty": 1, "UnitPrice": valor}
        if l.item_id:
            det["ItemRef"] = {"value": str(l.item_id)}
        linha = {"Amount": valor, "DetailType": "SalesItemLineDetail",
                 "SalesItemLineDetail": det}
        if l.descricao:
            linha["Description"] = l.descricao
        corpo["Line"].append(linha)

    if req.vence:
        corpo["DueDate"] = req.vence

    # mantém a convenção que já existe nas faturas de vocês
    memo = (req.memo or "").strip()
    if req.num and "estimate" not in memo.lower():
        marca = "Estimate #" + str(req.num).strip()
        memo = (marca + (" — " + memo if memo else ""))
    if memo:
        corpo["PrivateNote"] = memo

    j = qb_post("invoice", corpo)
    inv = j.get("Invoice") or {}
    _limpa_caches_fatura()
    log.info("Fatura %s criada no projeto %s por %s",
             inv.get("DocNumber"), req.qbid, quem or papel)
    return {"ok": True, "id": inv.get("Id"), "num": inv.get("DocNumber"),
            "total": float(inv.get("TotalAmt") or 0),
            "saldo": float(inv.get("Balance") or 0),
            "vence": inv.get("DueDate"), "data": inv.get("TxnDate")}


class EnvioReq(BaseModel):
    invoice_id: str = ""
    email: str = ""


@app.get("/qb/invoice/pdf")
def qb_invoice_pdf(invoice_id: str = "", num: str = "",
                   authorization: str = Header(default="")):
    """Devolve o PDF da fatura.

    Não sobe para o Storage de propósito: o bucket das Work Orders é
    público, e fatura tem valor e dado de cliente. Aqui o arquivo passa
    autenticado e vai direto para o aparelho de quem pediu.
    """
    quem_e(authorization)          # basta estar cadastrado (viewer também)
    if not invoice_id:
        raise HTTPException(400, "Sem a fatura.")
    dados = qb_get_pdf("invoice/" + str(invoice_id).strip() + "/pdf")
    nome = "Invoice_" + (_re.sub(r"[^A-Za-z0-9_-]+", "", str(num or "")) or
                         str(invoice_id)) + ".pdf"
    return StreamingResponse(
        io.BytesIO(dados), media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'})


@app.post("/qb/invoice/enviar")
def qb_invoice_enviar(req: EnvioReq, authorization: str = Header(default="")):
    """Manda a fatura por e-mail pelo próprio QuickBooks."""
    papel, quem = _pode_faturar(authorization)
    if not req.invoice_id:
        raise HTTPException(400, "Sem a fatura.")
    params = {}
    destino = (req.email or "").strip()
    if destino:
        params["sendTo"] = destino
    j = qb_post("invoice/" + str(req.invoice_id).strip() + "/send",
                params=params, octeto=True)
    inv = j.get("Invoice") or {}
    log.info("Fatura %s enviada para %s por %s",
             inv.get("DocNumber"), destino or "e-mail do cadastro",
             quem or papel)
    return {"ok": True, "num": inv.get("DocNumber"),
            "enviado_para": destino or
                            ((inv.get("BillEmail") or {}).get("Address") or ""),
            "status": inv.get("EmailStatus", "")}


# =====================================================================
# RECEBIDO NO PERÍODO — alimenta a tabela de metas por PM
#
# Por que isto é necessário: o /qb/abertas só enxerga fatura EM ABERTO.
# Quando o cliente paga, o projeto some daquela lista — então não dá para
# saber por ali quanto entrou no mês. Os Payment do QuickBooks são o
# registro da baixa, e é isso que esta rota lê.
# =====================================================================

_inv_cache = {}       # Id da fatura -> texto e classificação

# O app já classifica faturas em tipoFatura(): procura "down", "middle",
# "final", "full" no memo. Esta função segue o MESMO vocabulário — se as
# duas divergirem, a tela mostra um tipo e o total soma outro, e ninguém
# consegue conferir nada.
PALAVRAS_DOWN = ("down", "deposit")
PALAVRAS_CONTA = ("middle", "final", "full")


def _texto_fatura(inv):
    """Junta tudo o que pode conter a descrição, em minúsculas.

    O texto vive em lugares diferentes conforme quem criou a fatura
    (SumoQuote, QuickBooks à mão, importação), então olhamos todos.
    """
    partes = [
        str(inv.get("DocNumber") or ""),
        str(inv.get("PrivateNote") or ""),
        str((inv.get("CustomerMemo") or {}).get("value") or ""),
    ]
    for ln in (inv.get("Line") or []):
        partes.append(str(ln.get("Description") or ""))
        det = ln.get("SalesItemLineDetail") or {}
        partes.append(str((det.get("ItemRef") or {}).get("name") or ""))
    return " | ".join(p for p in partes if p).lower()


def _eh_entrada(texto):
    """Entrada não conta para o PM; middle, final e full contam.

    Middle e final ganham de "down" de propósito: uma fatura escrita
    "middle payment - down the hall bathroom" é um middle. O caso
    contrário (uma entrada que mencione "final") não aparece na prática.
    """
    if any(p in texto for p in PALAVRAS_CONTA):
        return False
    return any(p in texto for p in PALAVRAS_DOWN)


def qb_faturas_por_id(ids):
    """Busca faturas pelo Id, em lotes, guardando o que já foi lido."""
    faltam = [i for i in ids if i and i not in _inv_cache]
    for k in range(0, len(faltam), 40):
        lote = faltam[k:k + 40]
        lista = "','".join(_esc_sql(str(x)) for x in lote)
        try:
            achadas = qb_query(
                "select * from Invoice where Id in ('" + lista + "')"
            ).get("Invoice") or []
        except Exception:
            achadas = []
        for inv in achadas:
            txt = _texto_fatura(inv)
            _inv_cache[str(inv.get("Id"))] = {
                "texto": txt,
                "entrada": _eh_entrada(txt),
                "doc": inv.get("DocNumber") or "",
            }
        # Fatura que o QuickBooks não devolveu fica marcada como
        # DESCONHECIDA — e desconhecida não é o mesmo que "não é
        # entrada". Antes eu punha entrada=False e o valor entrava no
        # total do PM em silêncio: uma falha de leitura virava dinheiro.
        for x in lote:
            _inv_cache.setdefault(str(x), {"texto": "", "entrada": False,
                                           "doc": "", "desconhecida": True})
    return _inv_cache


def _linhas_do_pagamento(p):
    """Devolve (id_da_fatura, valor) para cada parte do pagamento.

    Um pagamento pode quitar mais de uma fatura de uma vez: nesse caso
    cada parte é classificada por conta própria.
    """
    saida = []
    for ln in (p.get("Line") or []):
        val = float(ln.get("Amount") or 0)
        if val <= 0:
            continue
        alvo = ""
        for lt in (ln.get("LinkedTxn") or []):
            if (lt.get("TxnType") or "") == "Invoice":
                alvo = str(lt.get("TxnId") or "")
                break
        saida.append((alvo, val))
    return saida


def qb_recebido_periodo(ini, fim, detalhado=False):
    """Soma, por projeto, o dinheiro que entrou entre duas datas.

    Só conta middle e final. Entradas (down payment) ficam de fora do
    total do PM, mesmo quando divididas em várias parcelas — a decisão
    é tomada na FATURA, então todas as partes dela caem junto.
    """
    chave = "v2:" + str(ini) + ".." + str(fim)      # v2: regra nova
    guardado = _rec_cache.get(chave)
    if guardado and (time.time() - guardado["quando"]) < 600 and not detalhado:
        return guardado["dados"]

    todos, pos = [], 1
    for _ in range(6):                      # até 6 mil baixas no período
        q = ("select * from Payment "
             "where TxnDate >= '" + _esc_sql(ini) + "' "
             "and TxnDate <= '" + _esc_sql(fim) + "' "
             "startposition " + str(pos) + " maxresults 1000")
        lote = qb_query(q).get("Payment") or []
        todos += lote
        if len(lote) < 1000:
            break
        pos += 1000

    # Uma consulta só para todas as faturas envolvidas, em lotes.
    ids = set()
    for p in todos:
        for fid, _v in _linhas_do_pagamento(p):
            if fid:
                ids.add(fid)
    qb_faturas_por_id(sorted(ids))

    por, linhas = {}, []
    for p in todos:
        ref = p.get("CustomerRef") or {}
        nome = ref.get("name") or ""
        if not nome:
            continue
        d = por.setdefault(nome, {"projeto": nome,
                                  "id": ref.get("value") or "",
                                  "recebido": 0.0, "qtd": 0,
                                  "entrada": 0.0, "semfatura": 0.0,
                                  "desconhecido": 0.0})
        partes = _linhas_do_pagamento(p)
        # Adiantamento sem fatura amarrada: entrou no caixa mas não
        # abateu nada. Nunca contou e continua não contando.
        naoaplicado = float(p.get("UnappliedAmt") or 0)
        if not partes:
            continue
        for fid, val in partes:
            inf = _inv_cache.get(fid) or {}
            if not fid:
                # parte sem fatura ligada: é adiantamento solto
                d["semfatura"] += val
                classe = "sem fatura"
            elif inf.get("desconhecida"):
                # Não consegui ler a fatura: fica de lado, visível, em
                # vez de entrar calado no total de alguém.
                d["desconhecido"] = d.get("desconhecido", 0.0) + val
                classe = "fatura não lida"
            elif inf.get("entrada"):
                d["entrada"] += val
                classe = "down payment"
            else:
                d["recebido"] += val
                d["qtd"] += 1
                classe = "conta"
            if detalhado:
                linhas.append({
                    "projeto": nome,
                    "data": p.get("TxnDate") or "",
                    "pagamento": p.get("Id") or "",
                    "fatura": inf.get("doc") or fid,
                    "valor": round(val, 2),
                    "classe": classe,
                    "texto": (inf.get("texto") or "")[:160],
                })
        if naoaplicado > 0 and detalhado:
            linhas.append({"projeto": nome, "data": p.get("TxnDate") or "",
                           "pagamento": p.get("Id") or "", "fatura": "",
                           "valor": round(naoaplicado, 2),
                           "classe": "não aplicado", "texto": ""})

    lista = sorted(por.values(), key=lambda x: -x["recebido"])
    for d in lista:
        d["recebido"] = round(d["recebido"], 2)
        d["entrada"] = round(d["entrada"], 2)
        d["semfatura"] = round(d["semfatura"], 2)
        d["desconhecido"] = round(d.get("desconhecido", 0.0), 2)
    lista = [d for d in lista
             if d["recebido"] or d["entrada"] or d["semfatura"]
             or d["desconhecido"]]
    if not detalhado:
        _rec_cache[chave] = {"quando": time.time(), "dados": lista}
    return (lista, linhas) if detalhado else lista


@app.get("/qb/recebido")
def qb_recebido(ini: str = "", fim: str = "",
                authorization: str = Header(default="")):
    """Quanto cada projeto pagou no período. Usado nas metas por PM.

    Sem datas, assume do dia 1º do mês corrente até hoje.
    """
    quem_e(authorization)                   # basta estar cadastrado no schedule
    hoje = _date.today()
    if not ini:
        ini = hoje.replace(day=1).isoformat()
    if not fim:
        fim = hoje.isoformat()
    lista = qb_recebido_periodo(ini, fim)
    return {"ok": True, "ini": ini, "fim": fim,
            "total": round(sum(d["recebido"] for d in lista), 2),
            "entradas": round(sum(d.get("entrada", 0) for d in lista), 2),
            "nao_lidas": round(sum(d.get("desconhecido", 0) for d in lista), 2),
            "projetos": lista}


@app.get("/qb/recebido/cliente")
def qb_recebido_cliente(nome: str = "", ini: str = "", fim: str = "",
                        authorization: str = Header(default="")):
    """Explica, para UM cliente, o que contou e o que não contou.

    Feito para responder à pergunta "por que este down payment entrou no
    total do PM?" sem ter que ler o relatório inteiro. Mostra a fatura,
    o texto em que a decisão se baseou e a classificação.
    """
    quem_e(authorization)
    hoje = _date.today()
    if not ini:
        ini = hoje.replace(day=1).isoformat()
    if not fim:
        fim = hoje.isoformat()
    _lista, linhas = qb_recebido_periodo(ini, fim, detalhado=True)
    alvo = (nome or "").strip().lower()
    if alvo:
        linhas = [l for l in linhas if alvo in (l["projeto"] or "").lower()]
    resumo = {}
    for l in linhas:
        resumo[l["classe"]] = round(resumo.get(l["classe"], 0) + l["valor"], 2)
    return {"ok": True, "cliente": nome, "ini": ini, "fim": fim,
            "palavras_entrada": list(PALAVRAS_DOWN),
            "palavras_que_contam": list(PALAVRAS_CONTA),
            "resumo_por_classe": resumo,
            "linhas": sorted(linhas, key=lambda x: x["data"])}


@app.get("/qb/recebido/detalhe")
def qb_recebido_detalhe(ini: str = "", fim: str = "",
                        authorization: str = Header(default="")):
    """Cada pagamento do período, a fatura que ele quitou e como foi
    classificado.

    Existe para conferir a regra contra o QuickBooks antes de confiar nos
    totais: eu não tenho como ver como as faturas estão escritas, então
    quem valida é quem olha. Se algum 'middle' aparecer como entrada, ou
    o contrário, a coluna 'texto' mostra em cima de qual palavra a
    decisão foi tomada.
    """
    quem_e(authorization)
    hoje = _date.today()
    if not ini:
        ini = hoje.replace(day=1).isoformat()
    if not fim:
        fim = hoje.isoformat()
    lista, linhas = qb_recebido_periodo(ini, fim, detalhado=True)
    resumo = {}
    for l in linhas:
        resumo[l["classe"]] = round(resumo.get(l["classe"], 0) + l["valor"], 2)
    return {"ok": True, "ini": ini, "fim": fim,
            "palavras_entrada": list(PALAVRAS_DOWN),
            "resumo_por_classe": resumo,
            "conta_para_o_pm": round(sum(d["recebido"] for d in lista), 2),
            "linhas": sorted(linhas, key=lambda x: (x["projeto"], x["data"]))}
